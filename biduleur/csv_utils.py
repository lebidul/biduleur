import os
import re
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any

import pandas as pd
import openpyxl

from biduleur.constants import (
    DATE, DATE_FIN, HORAIRE, FESTIVAL, STYLE_FESTIVAL, VILLE, LIEU, PRIX,
    COLONNE_INFO, detect_spectacle_columns,
)
from biduleur.event_utils import parse_bidul_event

import logging  # Ajouter cet import

log = logging.getLogger(__name__)  # Obtenir le logger pour ce module

# -----------------------------
# Helpers format PRIX
# -----------------------------

_CURRENCY_MAP = {
    "€": "€", "eur": "€", "euro": "€",
    "$": "$", "usd": "$", "dollar": "$",
    "£": "£", "gbp": "£", "pound": "£",
    "chf": "CHF", "franc": "CHF",
}

_CURRENCY_COL_CANDIDATES = ("DEVISE", "CURRENCY", "MONNAIE")

# Nom de la colonne optionnelle pour désactiver des lignes
INACTIF_COLUMN = "INACTIF"

# Mapping mois français
MOIS_FR = {
    1: 'janvier', 2: 'février', 3: 'mars', 4: 'avril',
    5: 'mai', 6: 'juin', 7: 'juillet', 8: 'août',
    9: 'septembre', 10: 'octobre', 11: 'novembre', 12: 'décembre'
}

# Jours de la semaine complets
JOURS_SEMAINE = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']

REQUIRED_BASE_COLUMNS = [
    DATE, HORAIRE, FESTIVAL, STYLE_FESTIVAL, VILLE, LIEU, PRIX,
]


def _canonical_currency_symbol(raw: Any) -> str:
    """
    Retourne un symbole canonique à partir d'un libellé devise (ex: 'EUR' -> '€').
    Par défaut retourne '€' si rien d'exploitable n'est fourni.
    """
    if raw is None:
        return "€"
    s = str(raw).strip().lower()
    return _CURRENCY_MAP.get(s, "€")


def _format_number_as_text_with_symbol(v: float, symbol: str) -> str:
    """
    Formate un nombre en texte + symbole. Supprime '.0' si entier.
    Exemple: 12.0 -> '12€', 12.5 -> '12.5€'
    (Si tu veux la virgule française, remplace le point par une virgule ici.)
    """
    try:
        f = float(v)
    except Exception:
        # Si non convertible, on le renvoie en str (dernier recours)
        return f"{v}"

    # entier ?
    if abs(f - round(f)) < 1e-9:
        return f"{int(round(f))}{symbol}"

    # sinon, on garde les décimales (sans zéros inutiles)
    txt = f"{f:.2f}".rstrip("0").rstrip(".")
    return f"{txt}{symbol}"


def _normalize_price_cell(value: Any, symbol_hint: str) -> str:
    """
    Normalise UNE cellule de PRIX en texte.
      - Si déjà string et contient un symbole monétaire, on renvoie tel quel (trim).
      - Si vide/None -> ''.
      - Si numérique -> format 'nombre + symbole' (par défaut, symbol_hint).
      - Si string '10' -> convertie en number -> '10€' etc.
    """
    if value is None:
        return ""

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return ""
        # Si déjà avec symbole (€ $ £ CHF) -> on préserve
        lowered = s.lower()
        if ("€" in s) or ("$" in s) or ("£" in s) or ("chf" in lowered) or ("eur" in lowered) or ("usd" in lowered) or (
                "gbp" in lowered):
            return s
        # Si c'est un nombre en texte -> on tente de convertir
        try:
            f = float(s.replace(",", "."))
            return _format_number_as_text_with_symbol(f, symbol_hint)
        except Exception:
            return s  # on laisse tel quel si non numérique

    # Numérique (int/float/Decimal)
    try:
        return _format_number_as_text_with_symbol(value, symbol_hint)
    except Exception:
        return f"{value}"


def _convert_price_column_to_text(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convertit la colonne PRIX (si présente) en texte, avec symbole de devise :
      - devise par défaut : €,
      - sinon, si une colonne 'DEVISE' / 'CURRENCY' / 'MONNAIE' est présente,
        on l'utilise pour déterminer un symbole par ligne.
    """
    if "PRIX" not in df.columns:
        return df

    # Détecte une éventuelle colonne de devise
    currency_col: Optional[str] = next((c for c in _CURRENCY_COL_CANDIDATES if c in df.columns), None)

    if currency_col:
        # Ligne à ligne, en tenant compte de la devise
        def _row_fmt(row: pd.Series) -> str:
            symbol = _canonical_currency_symbol(row.get(currency_col))
            return _normalize_price_cell(row["PRIX"], symbol)

        df["PRIX"] = df.apply(_row_fmt, axis=1)
    else:
        # Sans colonne devise : on force '€'
        df["PRIX"] = df["PRIX"].apply(lambda v: _normalize_price_cell(v, "€"))

    # S'assurer que c'est bien du texte (objet/str)
    df["PRIX"] = df["PRIX"].astype(str).fillna("")

    return df


def _filter_inactive_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filtre les lignes où la colonne INACTIF vaut "n" (case insensitive).
    La colonne INACTIF est optionnelle - si elle n'existe pas, aucun filtrage n'est fait.

    Args:
        df: DataFrame à filtrer

    Returns:
        DataFrame filtré (sans les lignes inactives)
    """
    # Chercher la colonne INACTIF (case insensitive)
    inactif_col = None
    for col in df.columns:
        if col.upper() == INACTIF_COLUMN:
            inactif_col = col
            break

    if inactif_col is None:
        # Colonne non présente, pas de filtrage
        return df

    # Compter les lignes avant filtrage
    count_before = len(df)

    # Filtrer : garder les lignes où INACTIF n'est PAS "o" (case insensitive)
    # On convertit en string, strip, et compare en lowercase
    mask = df[inactif_col].apply(
        lambda x: str(x).strip().lower() != "o" if pd.notna(x) else True
    )
    df_filtered = df[mask]

    # Compter les lignes filtrées
    count_filtered = count_before - len(df_filtered)
    if count_filtered > 0:
        log.info(f"[INACTIF] {count_filtered} ligne(s) ignorée(s) (colonne '{inactif_col}' = 'o')")

    return df_filtered


# -----------------------------
# Parsing des dates
# -----------------------------

def _parse_date(date_str: Any) -> tuple[str, str]:
    """
    Parse une date et retourne (sort_key, display_date).

    Formats supportés:
    - ISO: "2014-08-31" -> ("2014-08-31", "Dimanche 31 Août 2014")
    - Année seule: "2021" -> ("2021", "2021")
    - Mois-Année: "08-2021" -> ("2021-08", "Août 2021")
    - Jour + numéro: "Samedi 3" -> ("03", "Samedi 3")
    - Jour + numéro + mois: "Lun 03 février" -> ("03", "Lun 03 février")

    Returns:
        (sort_key, display_date)
        - sort_key: clé de tri (date ISO complète, année-mois ou jour paddé)
        - display_date: date formatée pour affichage
    """
    if not date_str:
        return ('', '')

    # Gérer NaT (pandas Not-a-Time) comme une valeur manquante
    if date_str is pd.NaT or (hasattr(pd, 'isna') and not isinstance(date_str, str) and pd.isna(date_str)):
        return ('', '')

    # Gérer les objets datetime/Timestamp (ex: depuis Excel)
    if isinstance(date_str, (datetime, pd.Timestamp)):
        jour_semaine = JOURS_SEMAINE[date_str.weekday()]
        mois_nom = MOIS_FR[date_str.month].capitalize()
        display = f"{jour_semaine} {date_str.day} {mois_nom} {date_str.year}"
        sort_key = f"{date_str.year:04d}-{date_str.month:02d}-{date_str.day:02d}"
        return (sort_key, display)

    # Gérer les entiers et flottants (ex: année seule 2005 ou 2005.0 depuis Excel)
    if isinstance(date_str, (int, float)):
        if not pd.isna(date_str) and date_str == int(date_str):
            date_str = str(int(date_str))
        else:
            return ('', str(date_str))

    if not isinstance(date_str, str):
        return ('', str(date_str))

    date_str = str(date_str).strip()

    # Format année seule: "2021" -> rendu tel quel
    if re.match(r'^\d{4}$', date_str):
        return (f"{date_str}-00-00", date_str)

    # Format mois-année: "08-2021" ou "08/2021" -> "Août 2021"
    mm_yyyy_match = re.match(r'^(\d{1,2})[-/](\d{4})$', date_str)
    if mm_yyyy_match:
        month, year = int(mm_yyyy_match.group(1)), int(mm_yyyy_match.group(2))
        if 1 <= month <= 12:
            mois_nom = MOIS_FR[month].capitalize()
            display = f"{mois_nom} {year}"
            sort_key = f"{year:04d}-{month:02d}-00"
            return (sort_key, display)

    # Format ISO: 2014-08-31 ou 2014/08/31 (avec optionnel 00:00:00)
    iso_match = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?$', date_str)
    if iso_match:
        year, month, day = int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3))
        try:
            dt = datetime(year, month, day)
            jour_semaine = JOURS_SEMAINE[dt.weekday()]
            mois_nom = MOIS_FR[month].capitalize()
            # Format complet avec année: "Dimanche 31 Août 2014"
            display = f"{jour_semaine} {day} {mois_nom} {year}"
            # Clé de tri: date ISO pour tri chronologique
            sort_key = f"{year:04d}-{month:02d}-{day:02d}"
            return (sort_key, display)
        except ValueError:
            return ('', date_str)

    # Format existant: "Samedi 3" ou "Sam 03 février"
    parts = date_str.split()
    if len(parts) >= 2:
        try:
            day = int(parts[1])
            # Clé de tri: jour paddé (comportement existant)
            return (str(day).zfill(2), date_str)
        except ValueError:
            pass

    return ('', date_str)


# -----------------------------
# Extraction des hyperlinks Excel
# -----------------------------

def _extract_hyperlinks(filename: str, df: pd.DataFrame) -> Dict[tuple, str]:
    """
    Extrait les hyperlinks d'un fichier Excel via openpyxl.
    Retourne un dict {(row_idx, col_name): target_url}.
    row_idx est l'index 0-based du DataFrame (correspondant à la ligne Excel - 2).
    """
    hyperlinks = {}
    try:
        wb = openpyxl.load_workbook(filename, read_only=True)
        ws = wb.active

        # Mapper les indices de colonnes Excel aux noms de colonnes du DataFrame
        col_map = {}  # excel_col_idx (1-based) -> df_col_name
        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
            if cell.value and str(cell.value).strip() in df.columns:
                col_map[col_idx] = str(cell.value).strip()

        # Relire le fichier en mode non read_only pour accéder aux hyperlinks
        wb.close()
        wb = openpyxl.load_workbook(filename, read_only=False)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.hyperlink and cell.hyperlink.target:
                    col_name = col_map.get(cell.column)
                    if col_name:
                        df_row_idx = cell.row - 2  # Excel row 2 = DataFrame index 0
                        hyperlinks[(df_row_idx, col_name)] = cell.hyperlink.target

        wb.close()
        if hyperlinks:
            log.info(f"[HYPERLINKS] {len(hyperlinks)} hyperlink(s) extraits du fichier Excel")
    except Exception as e:
        log.warning(f"[HYPERLINKS] Impossible d'extraire les hyperlinks: {e}")

    return hyperlinks


# -----------------------------
# Lecture + tri
# -----------------------------

def _is_full_date_key(sort_key: str) -> bool:
    """Vérifie si une clé de tri est une date complète (YYYY-MM-DD, sans '00')."""
    if not isinstance(sort_key, str):
        return False
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', sort_key)
    if not m:
        return False
    year, month, day = m.group(1), m.group(2), m.group(3)
    return year != '0000' and month != '00' and day != '00'


def _format_composite_date(group: List[tuple]) -> str:
    """
    Construit l'affichage groupé à partir d'une liste de (sort_key, display, datetime).

    Exemples:
      2 dates même mois: "Samedi 23 & Dimanche 24 Août 2025"
      3 dates même mois: "Samedi 23, Dimanche 24 & Lundi 25 Août 2025"
      Changement de mois: "Dimanche 31 Août & Lundi 1 Septembre 2025"
    """
    parts = []
    for i, (_sk, _disp, dt) in enumerate(group):
        jour_semaine = JOURS_SEMAINE[dt.weekday()]
        mois_nom = MOIS_FR[dt.month].capitalize()
        is_last = (i == len(group) - 1)

        if is_last:
            # Dernier: toujours afficher jour + numéro + mois + année
            parts.append(f"{jour_semaine} {dt.day} {mois_nom} {dt.year}")
        else:
            next_dt = group[i + 1][2]
            # Afficher le mois si le mois change avec le suivant
            show_month = (next_dt.month != dt.month)
            if show_month:
                parts.append(f"{jour_semaine} {dt.day} {mois_nom}")
            else:
                parts.append(f"{jour_semaine} {dt.day}")

    if len(parts) == 2:
        return f"{parts[0]} & {parts[1]}"
    return ", ".join(parts[:-1]) + " & " + parts[-1]


def _group_consecutive_dates(df: pd.DataFrame) -> None:
    """
    Regroupe les événements de jours calendaires consécutifs sous une même date
    d'affichage composite. Modifie la colonne DATE du DataFrame in-place.

    L'ordre original des événements (trié par clé 'Day') est préservé : seule
    la valeur affichée de la date change.

    Ne regroupe que les dates complètes (YYYY-MM-DD sans composante '00').
    Les dates partielles (année seule, mois-année) et 'En Bref' ne sont pas
    affectées.
    """
    if 'Day' not in df.columns or DATE not in df.columns:
        return

    # Extraire la séquence unique (sort_key, display) dans l'ordre d'apparition
    seen = set()
    unique_sequence = []
    for sk, disp in zip(df['Day'].tolist(), df[DATE].tolist()):
        if sk in seen:
            continue
        seen.add(sk)
        unique_sequence.append((sk, disp))

    # Parcourir la séquence et détecter les runs de dates consécutives
    replacements = {}
    current_run = []  # list of (sort_key, display, datetime)

    def _flush_run():
        if len(current_run) >= 2:
            composite = _format_composite_date(current_run)
            for sk, disp, _dt in current_run:
                replacements[disp] = composite

    for sk, disp in unique_sequence:
        if not _is_full_date_key(sk):
            _flush_run()
            current_run.clear()
            continue
        try:
            dt = datetime.strptime(sk, "%Y-%m-%d")
        except ValueError:
            _flush_run()
            current_run.clear()
            continue

        if current_run:
            last_dt = current_run[-1][2]
            if dt - last_dt == timedelta(days=1):
                current_run.append((sk, disp, dt))
                continue
            _flush_run()
            current_run.clear()

        current_run.append((sk, disp, dt))

    _flush_run()

    if replacements:
        df[DATE] = df[DATE].apply(lambda d: replacements.get(d, d))
        log.info(f"[DATES] {len(replacements)} date(s) regroupée(s) en "
                 f"{len(set(replacements.values()))} groupe(s) de jours consécutifs")


def _format_du_au_range(debut_dt: datetime, fin_dt: datetime) -> str:
    """
    Construit une plage de dates au format "Du <début> au <fin>".

    Exemples:
      Même mois:       "Du Samedi 27 au Dimanche 28 Août 2016"
      Mois différents: "Du Dimanche 31 Août au Lundi 1 Septembre 2025"
      Années différentes: "Du Mardi 31 Décembre 2024 au Mercredi 1 Janvier 2025"
    """
    jour_s_d = JOURS_SEMAINE[debut_dt.weekday()]
    jour_s_f = JOURS_SEMAINE[fin_dt.weekday()]
    mois_d = MOIS_FR[debut_dt.month].capitalize()
    mois_f = MOIS_FR[fin_dt.month].capitalize()

    if debut_dt.year != fin_dt.year:
        # Années différentes : afficher tout deux fois
        debut_str = f"{jour_s_d} {debut_dt.day} {mois_d} {debut_dt.year}"
        fin_str = f"{jour_s_f} {fin_dt.day} {mois_f} {fin_dt.year}"
    elif debut_dt.month != fin_dt.month:
        # Même année, mois différents : mois au début, année à la fin
        debut_str = f"{jour_s_d} {debut_dt.day} {mois_d}"
        fin_str = f"{jour_s_f} {fin_dt.day} {mois_f} {fin_dt.year}"
    else:
        # Même mois : pas de mois au début
        debut_str = f"{jour_s_d} {debut_dt.day}"
        fin_str = f"{jour_s_f} {fin_dt.day} {mois_f} {fin_dt.year}"

    return f"Du {debut_str} au {fin_str}"


def _apply_date_range_display(df: pd.DataFrame) -> None:
    """
    Pour les lignes où la colonne DATE FIN est renseignée, remplace la valeur
    affichée de DATE par une plage au format "Du <début> au <fin>".

    Exemple : DATE=2016-08-27, DATE FIN=2016-08-28
              → "Du Samedi 27 au Dimanche 28 Août 2016"

    La clé de tri 'Day' reste celle de la date de DÉBUT — les événements
    avec plage se placent chronologiquement selon leur date de début.

    Ne fait rien si la colonne DATE FIN n'existe pas dans le DataFrame ou si
    l'une des deux dates (début/fin) n'est pas une date complète.
    """
    if DATE_FIN not in df.columns or 'Day' not in df.columns:
        return

    nb_ranges = 0
    for idx in df.index:
        fin_raw = df.at[idx, DATE_FIN]
        if fin_raw is None:
            continue
        if isinstance(fin_raw, float) and pd.isna(fin_raw):
            continue
        if isinstance(fin_raw, str) and not fin_raw.strip():
            continue

        debut_sort_key = df.at[idx, 'Day']
        fin_sort_key, _ = _parse_date(fin_raw)

        if not _is_full_date_key(debut_sort_key) or not _is_full_date_key(fin_sort_key):
            continue

        try:
            debut_dt = datetime.strptime(debut_sort_key, "%Y-%m-%d")
            fin_dt = datetime.strptime(fin_sort_key, "%Y-%m-%d")
        except ValueError:
            continue

        if fin_dt <= debut_dt:
            continue  # Plage invalide ou non utile

        df.at[idx, DATE] = _format_du_au_range(debut_dt, fin_dt)
        nb_ranges += 1

    if nb_ranges:
        log.info(f"[DATES] {nb_ranges} plage(s) DATE→DATE FIN formatée(s)")


def read_and_sort_file(filename: str, date_grouping_enabled: bool = False) -> Optional[List[Dict]]:
    """
    Lit et trie un fichier (CSV, XLS, ou XLSX).
    Args:
        filename (str): Chemin du fichier d'entrée.
    Returns:
        Optional[List[Dict]]: Liste des enregistrements triés ou None en cas d'erreur.
    """
    try:
        file_extension = os.path.splitext(filename)[1].lower()

        # Lecture du fichier avec gestion d'erreur spécifique
        try:
            if file_extension in ['.csv']:
                df = pd.read_csv(filename, encoding='utf8', keep_default_na=False, na_values=[''])
            elif file_extension in ['.xls', '.xlsx']:
                df = pd.read_excel(filename, keep_default_na=False, na_values=[''])
            else:
                raise ValueError(
                    f"Format de fichier non supporté : {file_extension}\n\nFormats acceptés : .csv, .xls, .xlsx")
        except Exception as read_error:
            # Transformer les erreurs de lecture en messages clairs
            error_msg = str(read_error).lower()
            if "format cannot be determined" in error_msg or "corrupt" in error_msg:
                raise ValueError(
                    f"Le fichier Excel est corrompu ou illisible.\n\n"
                    f"Solutions :\n"
                    f"• Ouvrir le fichier dans Excel et l'enregistrer à nouveau\n"
                    f"• Exporter en CSV depuis Excel\n"
                    f"• Vérifier que le fichier n'est pas vide"
                )
            elif "permission" in error_msg or "access" in error_msg:
                raise ValueError(
                    f"Impossible d'accéder au fichier.\n\n"
                    f"Assurez-vous que :\n"
                    f"• Le fichier n'est pas ouvert dans Excel\n"
                    f"• Vous avez les droits de lecture"
                )
            else:
                raise ValueError(f"Erreur lors de la lecture du fichier :\n\n{read_error}")

        # Validation des colonnes de base obligatoires
        missing_cols = [col for col in REQUIRED_BASE_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(
                f"Colonnes manquantes dans le fichier :\n\n" +
                "\n".join(f"• {col}" for col in missing_cols)
            )

        # Détection dynamique des colonnes spectacle (GENRE N, NOM SPECTACLE N, etc.)
        spectacle_col_sets = detect_spectacle_columns(df.columns)
        if not spectacle_col_sets:
            raise ValueError(
                "Aucune colonne spectacle détectée (GENRE N, NOM SPECTACLE N, etc.).\n\n"
                "Le fichier doit contenir au moins un set de colonnes spectacle."
            )
        log.info(f"[COLONNES] {len(spectacle_col_sets)} set(s) de colonnes spectacle détecté(s) : "
                 f"{[s['num'] for s in spectacle_col_sets]}")

        # --- Extraction des hyperlinks depuis Excel ---
        hyperlinks = {}
        if file_extension in ['.xls', '.xlsx']:
            hyperlinks = _extract_hyperlinks(filename, df)

        # --- Injecter les hyperlinks dans le DataFrame ---
        # Pour les cellules avec un hyperlink, formater en "url|display_text"
        if hyperlinks:
            for (row_idx, col_name), target_url in hyperlinks.items():
                if row_idx < len(df) and col_name in df.columns:
                    cell_value = df.at[df.index[row_idx], col_name]
                    display_text = str(cell_value).strip() if pd.notna(cell_value) and cell_value else target_url
                    df.at[df.index[row_idx], col_name] = f"{target_url}|{display_text}"

        # --- Filtrage des lignes inactives ---
        df = _filter_inactive_rows(df)

        # --- Conversion PRIX -> texte avec devise ---
        df = _convert_price_column_to_text(df)

        # Remplace NaN par None pour la suite
        df = df.where(pd.notnull(df), None)

        # Parser les dates: extraire clé de tri + normaliser pour affichage
        parsed = df[DATE].apply(_parse_date)
        df['Day'] = parsed.apply(lambda x: x[0])  # Clé de tri (date ISO ou jour)
        df[DATE] = parsed.apply(lambda x: x[1])   # Date formatée pour affichage

        # Si une colonne DATE FIN existe, construire des plages composites
        # pour les lignes où elle est renseignée
        _apply_date_range_display(df)

        # Tri personnalisé : 'En Bref' (COLONNE_INFO) passe à la fin
        first_genre_col = spectacle_col_sets[0]['genre']
        df['info_last'] = df[DATE].apply(lambda x: 0 if x == COLONNE_INFO else 1)
        df_sorted = df.sort_values(by=['info_last', 'Day', first_genre_col, HORAIRE])
        df_sorted = df_sorted.drop(columns=['info_last'])

        # Regroupement optionnel des dates consécutives (feature Teriaki)
        if date_grouping_enabled:
            _group_consecutive_dates(df_sorted)

        records = df_sorted.to_dict('records')
        # Attacher les infos de colonnes spectacle aux records pour parse_bidul_event
        for rec in records:
            rec['_spectacle_col_sets'] = spectacle_col_sets
        return records
    except ValueError:
        # Re-lever les erreurs de validation pour qu'elles remontent
        raise
    except Exception as e:
        log.error(f"Error sorting the file: {e}. Ensure each line has a defined date.")
        return None


# -----------------------------
# Pipeline haut niveau
# -----------------------------

def parse_bidul(
        filename: str,
        date_grouping_enabled: bool = False,
        festival_in_date_header: bool = False,
        bidul_label_enabled: bool = False,
) -> tuple:
    """
    Traite le fichier d'entrée (CSV, XLS, ou XLSX) et génère les données nécessaires.
    Args:
        filename (str): Chemin du fichier d'entrée.
        date_grouping_enabled (bool): Si True, regroupe les événements de jours
            calendaires consécutifs sous une même date composite (feature Teriaki).
        festival_in_date_header (bool): Si True, déplace le contenu de la colonne
            FESTOCHE/EVENEMENT dans l'en-tête de date (feature expérimentale Teriaki).
    Returns:
        tuple: (html_body_bidul, html_body_agenda, number_of_lines)
    """
    body_content = ''
    body_content_agenda = ''
    sorted_events = read_and_sort_file(filename, date_grouping_enabled=date_grouping_enabled)
    if sorted_events is None:
        return body_content, body_content_agenda, 0

    current_date = None
    number_of_lines = 0

    for row in sorted_events:
        # strip des chaînes
        row = {key: (value.strip() if isinstance(value, str) else value) for key, value in row.items()}

        formatted_line_bidul, formatted_line_agenda, formatted_line_post, current_date = parse_bidul_event(
            row, current_date,
            festival_in_date_header=festival_in_date_header,
            bidul_label_enabled=bidul_label_enabled,
        )
        body_content += formatted_line_bidul + "\n\n"
        body_content_agenda += formatted_line_agenda + "\n\n"
        number_of_lines += 1

    return body_content, body_content_agenda, number_of_lines
"""Tests pour le module d'export CSV/XLSX."""

import pytest
import tempfile
import sqlite3
from pathlib import Path
import sys

# Ajouter le chemin parent pour les imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from core.csv_exporter import (
    XLSX_EXPORT_COLUMNS,
    INTERNAL_TO_XLSX,
    format_date,
    write_csv,
    write_xlsx,
    export_events,
    export_bidul,
)


class TestFormatDate:
    """Tests pour le formatage des dates."""

    def test_format_date_valid(self):
        """Test formatage d'une date valide."""
        assert format_date('2023-01-15') == 'Dimanche 15'
        assert format_date('2023-01-16') == 'Lundi 16'
        assert format_date('2023-01-17') == 'Mardi 17'

    def test_format_date_empty(self):
        """Test avec date vide."""
        assert format_date('') == ''
        assert format_date(None) == ''

    def test_format_date_invalid(self):
        """Test avec date invalide (retourne telle quelle)."""
        assert format_date('invalid') == 'invalid'
        assert format_date('2023/01/15') == '2023/01/15'


class TestColumnMapping:
    """Tests pour le mapping des colonnes."""

    def test_xlsx_columns_count(self):
        """Vérifie le nombre de colonnes XLSX."""
        assert len(XLSX_EXPORT_COLUMNS) == 23

    def test_internal_to_xlsx_mapping(self):
        """Vérifie que le mapping interne couvre les colonnes essentielles."""
        essential_keys = [
            'festival', 'style_festival', 'date', 'horaire',
            'lieu', 'ville', 'prix', 'genre1', 'spectacle1',
            'artiste1', 'style1'
        ]
        for key in essential_keys:
            assert key in INTERNAL_TO_XLSX, f"Clé manquante: {key}"

    def test_mapping_values_in_columns(self):
        """Vérifie que toutes les valeurs du mapping sont dans XLSX_EXPORT_COLUMNS."""
        for key, col_name in INTERNAL_TO_XLSX.items():
            assert col_name in XLSX_EXPORT_COLUMNS, \
                f"Colonne '{col_name}' pour '{key}' non trouvée dans XLSX_EXPORT_COLUMNS"


class TestWriteCsv:
    """Tests pour l'écriture CSV."""

    def test_write_csv_empty(self):
        """Test écriture CSV avec liste vide."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.csv'
            count = write_csv([], output_path)
            assert count == 0
            assert output_path.exists()

    def test_write_csv_single_row(self):
        """Test écriture CSV avec une seule ligne."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.csv'
            rows = [{
                INTERNAL_TO_XLSX['festival']: 'Test Festival',
                INTERNAL_TO_XLSX['date']: 'Lundi 1',
                INTERNAL_TO_XLSX['ville']: 'Le Mans',
            }]
            count = write_csv(rows, output_path)
            assert count == 1
            assert output_path.exists()

            # Vérifier le contenu
            content = output_path.read_text(encoding='utf-8')
            assert 'Test Festival' in content
            assert 'Le Mans' in content

    def test_write_csv_creates_parent_dirs(self):
        """Test que write_csv crée les dossiers parents."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'subdir' / 'deep' / 'test.csv'
            rows = [{INTERNAL_TO_XLSX['ville']: 'Test'}]
            count = write_csv(rows, output_path)
            assert count == 1
            assert output_path.exists()


class TestWriteXlsx:
    """Tests pour l'écriture XLSX."""

    def test_write_xlsx_empty(self):
        """Test écriture XLSX avec liste vide."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.xlsx'
            count = write_xlsx([], output_path)
            assert count == 0
            assert output_path.exists()

    def test_write_xlsx_single_row(self):
        """Test écriture XLSX avec une seule ligne."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.xlsx'
            rows = [{
                INTERNAL_TO_XLSX['festival']: 'Test Festival',
                INTERNAL_TO_XLSX['date']: 'Lundi 1',
                INTERNAL_TO_XLSX['ville']: 'Le Mans',
            }]
            count = write_xlsx(rows, output_path)
            assert count == 1
            assert output_path.exists()

    def test_write_xlsx_readable_by_openpyxl(self):
        """Test que le fichier XLSX est lisible par openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl non installé")

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.xlsx'
            rows = [{
                INTERNAL_TO_XLSX['festival']: 'Test Festival',
                INTERNAL_TO_XLSX['date']: 'Mardi 15',
                INTERNAL_TO_XLSX['ville']: 'Le Mans',
                INTERNAL_TO_XLSX['lieu']: 'La Fonderie',
            }]
            write_xlsx(rows, output_path)

            # Charger et vérifier
            wb = load_workbook(output_path)
            ws = wb.active
            assert ws.max_row == 2  # header + 1 row
            assert ws.max_column == 23

            # Vérifier le header
            assert ws.cell(row=1, column=1).value == XLSX_EXPORT_COLUMNS[0]

            # Vérifier les données
            assert ws.cell(row=2, column=1).value == 'Test Festival'

    def test_write_xlsx_preserves_newlines_in_headers(self):
        """Test que les newlines dans les headers sont préservés."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl non installé")

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test.xlsx'
            rows = [{INTERNAL_TO_XLSX['festival']: 'Test'}]
            write_xlsx(rows, output_path)

            wb = load_workbook(output_path)
            ws = wb.active
            # La première colonne contient un newline
            header = ws.cell(row=1, column=1).value
            assert '\n' in header, f"Header devrait contenir newline: {repr(header)}"


class TestExportEvents:
    """Tests pour l'export d'événements depuis la base."""

    @pytest.fixture
    def temp_db(self):
        """Crée une base de données temporaire avec des données de test."""
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / 'test.db'
            conn = sqlite3.connect(str(db_path))

            # Créer les tables
            conn.execute('''
                CREATE TABLE evenement (
                    id INTEGER PRIMARY KEY,
                    bidul_numero INTEGER,
                    date_evenement TEXT,
                    heure TEXT,
                    lieu_raw TEXT,
                    ville_raw TEXT,
                    tarif_raw TEXT,
                    nom TEXT,
                    genre_evenement TEXT,
                    type_evenement TEXT
                )
            ''')

            conn.execute('''
                CREATE TABLE contenu_evenement (
                    id INTEGER PRIMARY KEY,
                    evenement_id INTEGER,
                    nom_spectacle TEXT,
                    artiste TEXT,
                    style TEXT,
                    ordre INTEGER
                )
            ''')

            # Insérer des données de test
            conn.execute('''
                INSERT INTO evenement VALUES
                (1, 280, '2023-01-15', '20h30', 'La Fonderie', 'Le Mans', '10€',
                 'Festival Test', 'Rock', 'concert')
            ''')

            conn.execute('''
                INSERT INTO contenu_evenement VALUES
                (1, 1, 'Show 1', 'Artiste A', 'rock', 1),
                (2, 1, 'Show 2', 'Artiste B', 'jazz', 2)
            ''')

            conn.commit()
            conn.close()

            yield db_path

    def test_export_events_csv(self, temp_db):
        """Test export vers CSV."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'export.csv'
            count = export_events(temp_db, output_path, output_format='csv')
            assert count == 1
            assert output_path.exists()

            content = output_path.read_text(encoding='utf-8')
            assert 'Festival Test' in content
            assert 'Le Mans' in content

    def test_export_events_xlsx(self, temp_db):
        """Test export vers XLSX."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'export.xlsx'
            count = export_events(temp_db, output_path, output_format='xlsx')
            assert count == 1
            assert output_path.exists()

    def test_export_events_with_where_clause(self, temp_db):
        """Test export avec clause WHERE."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Ajouter un événement d'un autre bidul
            conn = sqlite3.connect(str(temp_db))
            conn.execute('''
                INSERT INTO evenement VALUES
                (2, 281, '2023-02-01', '21h', 'Bar XYZ', 'La Flèche', '5€',
                 'Concert Solo', 'Jazz', 'concert')
            ''')
            conn.commit()
            conn.close()

            # Exporter seulement bidul 280
            output_path = Path(tmpdir) / 'export.csv'
            count = export_events(
                temp_db,
                output_path,
                where_clause='e.bidul_numero = ?',
                params=(280,)
            )
            assert count == 1

            content = output_path.read_text(encoding='utf-8')
            assert 'Festival Test' in content
            assert 'Concert Solo' not in content

    def test_export_bidul(self, temp_db):
        """Test export d'un bidul spécifique."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'bidul_280.csv'
            count = export_bidul(temp_db, 280, output_path)
            assert count == 1

    def test_export_no_results(self, temp_db):
        """Test export sans résultats."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'export.csv'
            count = export_events(
                temp_db,
                output_path,
                where_clause='e.bidul_numero = ?',
                params=(999,)  # N'existe pas
            )
            assert count == 0
